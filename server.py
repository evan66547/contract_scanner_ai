import os
import json
import base64
import asyncio
import hashlib
import hmac
import importlib.util
import time
import urllib.parse
from contextlib import asynccontextmanager
from typing import Any, List, Optional, Set
from io import BytesIO, StringIO
import csv
import socket
import subprocess
import re

import multiprocessing
from ocr import OcrRuntime
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import openpyxl
import aiohttp
import logging
from logging.handlers import RotatingFileHandler

# === 服务配置 ===
SERVER_PORT = int(os.environ.get('PORT', os.environ.get('SERVER_PORT', 8093)))
DEBUG_WS = os.environ.get("DEBUG_WS", "0") == "1"

# === 全局 aiohttp Session（复用连接池，避免每次请求创建新连接） ===
http_session: Optional[aiohttp.ClientSession] = None

async def _reset_http_session():
    """重置 aiohttp session，清理可能卡死的连接池，同步更新 runtime 引用"""
    global http_session, _ocr_runtime, _fallback_ocr_runtime
    if http_session and not http_session.closed:
        await http_session.close()
    timeout = aiohttp.ClientTimeout(total=60, connect=10)
    http_session = aiohttp.ClientSession(timeout=timeout)
    for rt in (_ocr_runtime, _fallback_ocr_runtime):
        if rt is not None and not rt._closed:
            rt.http_session = http_session
    logger.info("🔄 aiohttp Session 已重置（清理死连接）")

# === WebSocket 并发控制 ===
MAX_WS_CONNECTIONS = 5  # Support up to 5 devices
MAX_OCR_CONCURRENT = 5  # 增加并发限制
ws_semaphore: Optional[asyncio.Semaphore] = None
active_ws_clients: Set[WebSocket] = set()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """管理应用生命周期：启动时创建 HTTP Session，预热模型，关闭时释放资源"""
    global http_session, ws_semaphore
    http_session = aiohttp.ClientSession(
        timeout=aiohttp.ClientTimeout(total=10, connect=2)
    )
    ws_semaphore = asyncio.Semaphore(MAX_OCR_CONCURRENT)
    logger.info("✅ aiohttp Session 已创建")
    logger.info(f"✅ OCR 并发限制: {MAX_OCR_CONCURRENT}")

    # 初始化 OCR Runtime（Phase 1: 配置管理 + mtime 缓存）
    global _ocr_runtime
    _ocr_runtime = OcrRuntime(CONFIG_FILE, http_session)
    app.state.ocr = _ocr_runtime

    # 预热 GLM-OCR 模型（加载到显存/内存）
    asyncio.create_task(warmup_model())

    yield
    global _fallback_ocr_runtime
    if _ocr_runtime is not None:
        await _ocr_runtime.close()
        logger.info("🔚 OCR Runtime 已关闭")
    if _fallback_ocr_runtime is not None:
        await _fallback_ocr_runtime.close()
        _fallback_ocr_runtime = None
        logger.info("🔚 Fallback OCR Runtime 已关闭")
    await http_session.close()
    logger.info("🔚 aiohttp Session 已关闭")

async def warmup_model():
    """启动后异步预热当前本地模型。"""
    try:
        cfg = get_ocr_config()
        provider = cfg["ocr"].get("provider", "mlx")
        if provider == "mlx":
            logger.info("⏭️ MLX GLM-OCR 跳过启动预热；首次扫描时尝试加载，失败则自动降级到 PaddleOCR")
            return
        if provider != "ollama":
            logger.info("⏭️ 当前非本地大模型引擎，跳过预热")
            return
        ollama_cfg = cfg["ocr"].get("ollama", {})
        base_url = ollama_cfg.get("baseUrl", "http://localhost:11434")
        model = ollama_cfg.get("model", "glm-ocr")

        # 检查当前是否已有其他模型占用显存
        async with http_session.get(f"{base_url}/api/ps", timeout=aiohttp.ClientTimeout(total=5)) as ps_resp:
            if ps_resp.status == 200:
                ps_data = await ps_resp.json()
                loaded = [m.get("name", "") for m in ps_data.get("models", [])]
                other_loaded = [n for n in loaded if model not in n]
                if other_loaded:
                    logger.info(f"⏭️ 检测到其他模型运行中 ({', '.join(other_loaded)})，跳过预热避免抢占显存")
                    return
                if any(model in n for n in loaded):
                    logger.info(f"✅ {model} 已在内存中，无需预热")
                    return

        logger.info(f"⏳ 正在预热模型 {model}...")
        url = f"{base_url.rstrip('/')}/api/generate"
        payload = {
            "model": model,
            "prompt": "hi",
            "stream": False,
            "keep_alive": ollama_cfg.get("keepAlive", "10m"),
            "options": {"num_predict": 1, "temperature": 0},
        }
        async with http_session.post(url, json=payload, timeout=aiohttp.ClientTimeout(total=120)) as resp:
            if resp.status == 200:
                logger.info(f"✅ 模型 {model} 预热完成，已加载到内存")
            else:
                logger.warning(f"⚠️ 模型预热失败: HTTP {resp.status}")
    except Exception as e:
        logger.warning(f"⚠️ 模型预热异常: {e}")

def _mlx_runtime_status(model_name: str = "mlx-community/GLM-OCR-8bit") -> tuple[bool, str]:
    if not importlib.util.find_spec("mlx_vlm"):
        return False, "未安装 mlx-vlm"
    if not OcrRuntime.is_mlx_model_cached(model_name):
        return False, f"模型权重未完整下载: {model_name}"
    try:
        import mlx.core as mx
        mx.metal.device_info()
        return True, "ready"
    except Exception as e:
        return False, str(e)

def _paddle_runtime_status() -> tuple[bool, str]:
    if not importlib.util.find_spec("paddleocr"):
        return False, "未安装 paddleocr"
    if not importlib.util.find_spec("paddle"):
        return False, "未安装 paddlepaddle"
    return True, "PaddleOCR 可用"

def _warmup_image_b64() -> str:
    """1x1 white JPEG for local model warmup."""
    from PIL import Image

    buf = BytesIO()
    Image.new("RGB", (1, 1), "white").save(buf, format="JPEG")
    return base64.b64encode(buf.getvalue()).decode("ascii")

app = FastAPI(title="Contract Scanner AI", lifespan=lifespan)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# === 日志系统 ===
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "scanner.log")

logger = logging.getLogger("scanner")
logger.setLevel(logging.INFO)
# Prevent duplicate handlers on reload
if not logger.handlers:
    handler = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] [%(module)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    logger.addHandler(handler)
    # Also log to console
    console = logging.StreamHandler()
    console.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"))
    logger.addHandler(console)

CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
TARGETS_FILE = os.path.join(BASE_DIR, "targets.json")

NAME_ALIASES = {'识别对象', '公司名称', '公司名', '名称', '公司', '目标', '对象', 'company', 'name', 'target', '企业名称', '单位名称', '客户名称'}
INFO_ALIASES = {'显示信息', '附加信息', '备注', '日期', '开单日期', '合同编号', 'info', 'note', 'date', '说明'}

# ADB 无线调试：缓存手机 WiFi IP（开启 tcpip 时记录，拔线后仍可用）
adb_wifi_ip_cache: Optional[str] = None
INFO2_ALIASES = {'合同总额', '欠款金额', '金额', '总额', 'amount', 'total', 'balance', 'debt', '合同金额', '订单金额'}


def _parse_adb_device_lines(output: str) -> list:
    devices = []
    for raw_line in output.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("List"):
            continue
        parts = line.split()
        if len(parts) < 2:
            continue
        serial = parts[0]
        adb_state = parts[1]
        model = ""
        for p in parts:
            if p.startswith("model:"):
                model = p.split(":", 1)[1]
                break
        is_network_serial = ":" in serial and not serial.startswith("emulator-")
        is_usb = not is_network_serial
        if "usb:" in line:
            is_usb = True
        elif "transport_id:" in line and not is_network_serial:
            is_usb = True
        devices.append({
            "serial": serial,
            "adb_state": adb_state,
            "mode": "usb" if is_usb else "wifi",
            "model": model,
            "raw": line,
        })
    return devices


def _get_usb_android_devices() -> list:
    """Best-effort macOS USB detection for cases where USB sees Pixel but ADB does not."""
    if os.uname().sysname != "Darwin":
        return []
    try:
        r = subprocess.run(["ioreg", "-p", "IOUSB", "-l", "-w", "0"],
                           capture_output=True, text=True, timeout=5)
    except Exception:
        return []
    if r.returncode != 0:
        return []

    devices = []
    blocks = re.split(r"\n\s*\+-o ", r.stdout)
    for block in blocks:
        if '"USB Vendor Name" = "Google"' not in block and '"kUSBVendorString" = "Google"' not in block:
            continue
        product_match = re.search(r'"USB Product Name" = "([^"]+)"', block)
        if not product_match:
            product_match = re.search(r'"kUSBProductString" = "([^"]+)"', block)
        serial_match = re.search(r'"USB Serial Number" = "([^"]+)"', block)
        if not serial_match:
            serial_match = re.search(r'"kUSBSerialNumberString" = "([^"]+)"', block)
        devices.append({
            "product": product_match.group(1) if product_match else "Google Android device",
            "serial": serial_match.group(1) if serial_match else "",
        })
    return devices


def _adb_error_message_from_devices(devices: list) -> str:
    if not devices:
        usb_devices = _get_usb_android_devices()
        if usb_devices:
            names = ", ".join(
                f"{d['product']} {d['serial']}".strip()
                for d in usb_devices
            )
            return f"macOS 已识别 {names}，但 ADB 未识别。请在 Pixel 开发者选项中开启 USB 调试；若已开启，请撤销 USB 调试授权后重新插线并允许这台电脑。"
        return "未检测到 Android 设备。请确认 Pixel 已插入、USB 线支持数据传输，并已开启 USB 调试。"
    unauthorized = [d["serial"] for d in devices if d["adb_state"] == "unauthorized"]
    if unauthorized:
        return "Pixel 未授权 USB 调试。请解锁手机，在弹窗中允许这台电脑调试，然后重试。"
    offline = [d["serial"] for d in devices if d["adb_state"] == "offline"]
    if offline:
        return "Pixel ADB 状态为 offline。请重新插拔 USB，必要时执行 adb kill-server 后重试。"
    statuses = ", ".join(f"{d['serial']}={d['adb_state']}" for d in devices)
    return f"未检测到可用 Android 设备: {statuses}"


def _adb_device_public_fields(devices: list) -> list:
    return [
        {
            "serial": d["serial"],
            "mode": d["mode"],
            "model": d["model"],
            "adb_state": d["adb_state"],
        }
        for d in devices
    ]

# Pydantic models
class ExcelUploadResult(BaseModel):
    headers: List[str]
    preview: List[dict]
    totalRows: int
    nameCol: int
    infoCol: int
    infoCol2: int
    allRows: List[List[Any]]

class ImportConfirmReq(BaseModel):
    nameCol: int
    infoCol: int
    infoCol2: int
    headers: List[str] = []
    allRows: List[List[Any]]

class AdbWifiConnectReq(BaseModel):
    wifi_ip: Optional[str] = None

def load_json(filepath, default_val):
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error loading {filepath}: {e}")
    return default_val

def save_json(filepath, data):
    import tempfile
    dir_name = os.path.dirname(filepath) or '.'
    with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.tmp', delete=False, dir=dir_name) as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(f.name, filepath)

_ocr_runtime: Optional[OcrRuntime] = None
_fallback_ocr_runtime: Optional[OcrRuntime] = None

def get_ocr_config():
    """获取 OCR 配置：优先走 OcrRuntime（mtime 缓存），降级走原逻辑"""
    if _ocr_runtime is not None:
        return _ocr_runtime.get_config()
    # 降级：lifespan 前或测试环境
    cfg = load_json(CONFIG_FILE, {})
    if "ocr" not in cfg:
        old_model = cfg.pop("ollamaModel", "glm-ocr")
        cfg["ocr"] = {
            "provider": "mlx",
            "mlx": {
                "model": "mlx-community/GLM-OCR-8bit",
                "maxTokens": 160,
                "temperature": 0.0,
                "prompt": OcrRuntime._default_ocr_prompt(),
            },
            "ollama": {"baseUrl": "http://localhost:11434", "model": old_model, "keepAlive": "10m"},
            "baidu": {"apiKey": "", "secretKey": ""},
            "ocrspace": {"apiKey": "", "language": "chs"},
            "openai": {"apiKey": "", "model": "gpt-4o-mini", "baseUrl": "https://api.openai.com/v1"},
            "paddle": {"useGpu": False}
        }
        save_json(CONFIG_FILE, cfg)
    ocr_cfg = cfg.setdefault("ocr", {})
    if "mlx" not in ocr_cfg:
        ocr_cfg["mlx"] = {
            "model": "mlx-community/GLM-OCR-8bit",
            "maxTokens": 160,
            "temperature": 0.0,
            "prompt": OcrRuntime._default_ocr_prompt(),
        }
    if "ocrspace" not in cfg.get("ocr", {}):
        cfg["ocr"]["ocrspace"] = {"apiKey": "", "language": "chs"}
    if "paddle" not in cfg.get("ocr", {}):
        cfg["ocr"]["paddle"] = {"useGpu": False}
    return cfg

def _runtime_for_ocr() -> OcrRuntime:
    """获取 OCR Runtime；优先用 lifespan 实例，否则返回单例 fallback。"""
    global _fallback_ocr_runtime
    if _ocr_runtime is not None:
        return _ocr_runtime
    if _fallback_ocr_runtime is None:
        _fallback_ocr_runtime = OcrRuntime(CONFIG_FILE, http_session)
    return _fallback_ocr_runtime

async def get_baidu_token(api_key: str, secret_key: str) -> str:
    runtime = _runtime_for_ocr()
    if runtime.http_session is None:
        return ""
    return await runtime.get_baidu_token(api_key, secret_key)

# ================================
# OCR 引擎调度器
# ================================
async def dispatch_ocr(base64_img: str) -> str:
    if _ocr_runtime is not None:
        return await _ocr_runtime.recognize_text(base64_img)
    # 降级：无 OcrRuntime 时保留原逻辑
    cfg = get_ocr_config()
    provider = cfg["ocr"].get("provider", "mlx")
    if provider == "mlx":
        return await call_mlx(base64_img, cfg["ocr"].get("mlx", {}))
    elif provider == "baidu":
        return await call_baidu(base64_img, cfg["ocr"]["baidu"])
    elif provider == "ocrspace":
        return await call_ocrspace(base64_img, cfg["ocr"].get("ocrspace", {}))
    elif provider == "openai":
        return await call_openai(base64_img, cfg["ocr"]["openai"])
    elif provider == "paddle":
        return await call_paddle(base64_img)
    else:
        return await call_ollama(base64_img, cfg["ocr"].get("ollama", {}))

async def call_ollama(base64_img: str, ollama_cfg: dict = None) -> str:
    if ollama_cfg is None:
        cfg = get_ocr_config()
        ollama_cfg = cfg["ocr"].get("ollama", {})
    return await _runtime_for_ocr().recognize_ollama(base64_img, ollama_cfg)

async def call_mlx(base64_img: str, mlx_cfg: dict = None) -> str:
    if mlx_cfg is None:
        cfg = get_ocr_config()
        mlx_cfg = cfg["ocr"].get("mlx", {})
    return await _runtime_for_ocr().recognize_mlx(base64_img, mlx_cfg)

async def call_baidu(base64_img: str, baidu_cfg: dict) -> str:
    return await _runtime_for_ocr().recognize_baidu(base64_img, baidu_cfg)

async def call_ocrspace(base64_img: str, ocrspace_cfg: dict) -> str:
    api_key = ocrspace_cfg.get("apiKey", "")
    if not api_key:
        return "[OCR.space Error: API Key 未配置]"
    lang = ocrspace_cfg.get("language", "chs")

    url = "https://api.ocr.space/parse/image"
    payload = aiohttp.FormData()
    payload.add_field("base64Image", f"data:image/jpeg;base64,{base64_img}")
    payload.add_field("language", lang)
    payload.add_field("isOverlayRequired", "false")

    headers = {"apikey": api_key}

    try:
        async with http_session.post(url, data=payload, headers=headers) as resp:
            data = await resp.json()
            if data.get("OCRExitCode") != 1:
                return f"[OCR.space Error: {data.get('ErrorMessage', 'unknown')}]"
            results = data.get("ParsedResults", [])
            if not results:
                return ""
            return results[0].get("ParsedText", "").replace("\r\n", "").replace("\n", "")
    except Exception as e:
        return f"[OCR.space Error: {str(e)}]"

async def call_openai(base64_img: str, openai_cfg: dict) -> str:
    api_key = openai_cfg.get("apiKey", "")
    if not api_key:
        return "[OpenAI Error: API Key 未配置]"
    base_url = openai_cfg.get("baseUrl", "https://api.openai.com/v1").rstrip("/")
    model = openai_cfg.get("model", "gpt-4o-mini")

    url = f"{base_url}/chat/completions"
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": [
            {"type": "text", "text": "提取图片中的所有文字，只输出文字内容，不加任何解释："},
            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_img}"}}
        ]}],
        "max_tokens": 1000
    }
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    try:
        async with http_session.post(url, json=payload, headers=headers) as resp:
            data = await resp.json()
            if "error" in data:
                return f"[OpenAI Error: {data['error'].get('message', 'unknown')}]"
            return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"[OpenAI Error: {str(e)}]"

async def call_paddle(base64_img: str) -> str:
    return await _runtime_for_ocr().recognize_paddle(base64_img)

@app.post("/api/ocr-test")
async def ocr_test():
    """测试当前 OCR 引擎连通性"""
    cfg = get_ocr_config()
    provider = cfg["ocr"].get("provider", "mlx")
    result = {"provider": provider, "ok": False, "message": ""}

    try:
        if provider == "mlx":
            mlx_cfg = cfg["ocr"].get("mlx", {})
            model_name = mlx_cfg.get("model", "mlx-community/GLM-OCR-8bit")
            mlx_ready, mlx_message = _mlx_runtime_status(model_name)
            if mlx_ready:
                result["ok"] = True
                result["message"] = f"MLX GLM-OCR 可用，模型: {model_name}"
            else:
                paddle_ready, paddle_message = _paddle_runtime_status()
                result["ok"] = paddle_ready
                result["message"] = (
                    f"MLX 未就绪，已降级到 PaddleOCR: {mlx_message}"
                    if paddle_ready
                    else f"MLX 不可用，且 PaddleOCR 不可用: {mlx_message}; {paddle_message}"
                )
        elif provider == "ollama":
            ollama_cfg = cfg["ocr"].get("ollama", {})
            base_url = ollama_cfg.get("baseUrl", "http://localhost:11434")
            async with http_session.get(f"{base_url}/api/tags", timeout=aiohttp.ClientTimeout(total=5)) as resp:
                if resp.status == 200:
                    result["ok"] = True
                    result["message"] = f"Ollama 在线，模型: {ollama_cfg.get('model', 'glm-ocr')}"
                else:
                    result["message"] = f"Ollama 返回 {resp.status}"
        elif provider == "baidu":
            baidu_cfg = cfg["ocr"].get("baidu", {})
            if not baidu_cfg.get("apiKey"):
                result["message"] = "API Key 未配置"
            else:
                token = await get_baidu_token(baidu_cfg["apiKey"], baidu_cfg.get("secretKey", ""))
                result["ok"] = bool(token)
                result["message"] = "百度 OCR 连接成功" if token else "获取 access_token 失败"
        elif provider == "ocrspace":
            ocrspace_cfg = cfg["ocr"].get("ocrspace", {})
            if not ocrspace_cfg.get("apiKey"):
                result["message"] = "API Key 未配置"
            else:
                result["ok"] = True
                result["message"] = "OCR.space 已配置"
        elif provider == "openai":
            openai_cfg = cfg["ocr"].get("openai", {})
            if not openai_cfg.get("apiKey"):
                result["message"] = "API Key 未配置"
            else:
                result["ok"] = True
                result["message"] = f"OpenAI 已配置，模型: {openai_cfg.get('model', 'gpt-4o-mini')}"
        elif provider == "paddle":
            try:
                import paddleocr
                result["ok"] = True
                result["message"] = f"PaddleOCR 已安装，版本: {paddleocr.__version__}"
            except ImportError:
                result["message"] = "PaddleOCR 未安装，请运行: pip install paddleocr"
    except Exception as e:
        result["message"] = str(e)

    return result

_SECRET_FIELDS = [
    ("ocr", "baidu", "apiKey"),
    ("ocr", "baidu", "secretKey"),
    ("ocr", "ocrspace", "apiKey"),
    ("ocr", "openai", "apiKey"),
]
_REDACTED = "••••••••"

# === 扫描计数（按日期持久化到 scan_stats.json） ===
SCAN_STATS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scan_stats.json")
_scan_stats_cache: Optional[dict] = None
_scan_stats_mtime: float = 0.0
_scan_stats_lock = asyncio.Lock()

def _today_str() -> str:
    from datetime import date
    return date.today().isoformat()

def _load_scan_stats() -> dict:
    """加载 scan_stats.json，mtime 变化时重新读取"""
    global _scan_stats_cache, _scan_stats_mtime
    try:
        current_mtime = os.stat(SCAN_STATS_FILE).st_mtime
    except OSError:
        current_mtime = 0.0
    if _scan_stats_cache is not None and current_mtime == _scan_stats_mtime:
        return _scan_stats_cache
    try:
        if os.path.exists(SCAN_STATS_FILE):
            with open(SCAN_STATS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                _scan_stats_cache = data
                _scan_stats_mtime = current_mtime
                return data
    except Exception:
        pass
    return {}

def _save_scan_stats(stats: dict):
    """原子写入 scan_stats.json，同步维护内存缓存"""
    global _scan_stats_cache, _scan_stats_mtime
    import tempfile
    dir_name = os.path.dirname(SCAN_STATS_FILE) or "."
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", suffix=".tmp", delete=False, dir=dir_name
    ) as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    os.replace(f.name, SCAN_STATS_FILE)
    _scan_stats_cache = stats
    try:
        _scan_stats_mtime = os.stat(SCAN_STATS_FILE).st_mtime
    except OSError:
        _scan_stats_mtime = 0.0

def _coerce_scan_count(value) -> int:
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return 0

def _is_date_key(value: str) -> bool:
    import re
    return isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", value) is not None

def _sanitize_device_id(value) -> str:
    import re
    raw = str(value or "").strip()
    if not raw:
        return ""
    return re.sub(r"[^A-Za-z0-9_.:-]+", "-", raw)[:80]

def _sanitize_device_label(value, fallback: str = "") -> str:
    raw = str(value or "").strip()
    if not raw:
        raw = fallback or "未知设备"
    return raw[:40]

def _normalize_scan_stats(stats: dict) -> dict:
    if not isinstance(stats, dict):
        stats = {}
    global_stats = stats.get("global") if isinstance(stats.get("global"), dict) else {}
    for key, value in list(stats.items()):
        if _is_date_key(key) and key not in global_stats:
            global_stats[key] = value
    devices = stats.get("devices") if isinstance(stats.get("devices"), dict) else {}
    device_meta = stats.get("deviceMeta") if isinstance(stats.get("deviceMeta"), dict) else {}
    stats["schema"] = "v2"
    stats["global"] = global_stats
    stats["devices"] = devices
    stats["deviceMeta"] = device_meta
    return stats

def _get_global_scan_count(stats: dict, target_date: str) -> int:
    stats = _normalize_scan_stats(stats)
    return _coerce_scan_count(stats["global"].get(target_date, stats.get(target_date, 0)))

def _set_global_scan_count(stats: dict, target_date: str, count: int):
    count = _coerce_scan_count(count)
    stats = _normalize_scan_stats(stats)
    stats["global"][target_date] = count
    # Keep legacy top-level date mirror so older local scripts still read totals.
    stats[target_date] = count

def _get_device_scan_count(stats: dict, device_id: str, target_date: str) -> int:
    stats = _normalize_scan_stats(stats)
    device_stats = stats["devices"].get(device_id)
    if not isinstance(device_stats, dict):
        return 0
    return _coerce_scan_count(device_stats.get(target_date, 0))

def _set_device_scan_count(stats: dict, device_id: str, target_date: str, count: int, label: str = ""):
    device_id = _sanitize_device_id(device_id)
    if not device_id:
        return
    stats = _normalize_scan_stats(stats)
    device_stats = stats["devices"].setdefault(device_id, {})
    device_stats[target_date] = _coerce_scan_count(count)
    meta = stats["deviceMeta"].setdefault(device_id, {})
    if label:
        meta["label"] = _sanitize_device_label(label, device_id)
    meta["lastSeen"] = target_date

def _device_scan_rows(stats: dict, target_date: str) -> list:
    stats = _normalize_scan_stats(stats)
    rows = []
    device_total = 0
    for device_id, device_stats in stats["devices"].items():
        if not isinstance(device_stats, dict):
            continue
        count = _coerce_scan_count(device_stats.get(target_date, 0))
        device_total += count
        meta = stats["deviceMeta"].get(device_id, {}) if isinstance(stats["deviceMeta"].get(device_id), dict) else {}
        label = _sanitize_device_label(meta.get("label"), device_id)
        if label == "K" and device_id.startswith("dev-"):
            label = "Android 手机"
        rows.append({
            "device_id": device_id,
            "label": label,
            "count": count,
            "lastSeen": meta.get("lastSeen", "")
        })
    unassigned = max(0, _get_global_scan_count(stats, target_date) - device_total)
    if unassigned > 0:
        rows.append({
            "device_id": "unassigned",
            "label": "未归属设备",
            "count": unassigned,
            "lastSeen": target_date
        })
    rows.sort(key=lambda item: (-item["count"], item["label"]))
    return rows

def _is_error_text(text: str) -> bool:
    """判断 OCR 结果是否为错误/超时字符串"""
    if not text:
        return True
    if text.startswith("[") and ("Error:" in text or "Timeout:" in text):
        return True
    return False

async def _maybe_count_scan(text: str, device_id: str = "", device_label: str = "") -> Optional[int]:
    """OCR 结果满足条件时计数 +1：去空白后长度 > 6 且非错误字符串"""
    if _is_error_text(text):
        return None
    cleaned = text.replace("\n", "").replace("\r", "").replace(" ", "")
    if len(cleaned) > 6:
        async with _scan_stats_lock:
            global _scan_stats_cache
            today = _today_str()
            stats = _load_scan_stats()
            total = _get_global_scan_count(stats, today) + 1
            _set_global_scan_count(stats, today, total)
            clean_device_id = _sanitize_device_id(device_id)
            if clean_device_id:
                device_count = _get_device_scan_count(stats, clean_device_id, today) + 1
                _set_device_scan_count(stats, clean_device_id, today, device_count, device_label)
            _save_scan_stats(stats)
            _scan_stats_cache = stats
            return total
    return None

@app.get("/api/scan-stats")
async def get_scan_stats():
    today = _today_str()
    stats = _load_scan_stats()
    return {
        "date": today,
        "count": _get_global_scan_count(stats, today),
        "devices": _device_scan_rows(stats, today)
    }

@app.get("/api/scan-stats/month")
async def get_scan_stats_month(month: str = ""):
    import re, calendar
    from datetime import date
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        month = date.today().strftime("%Y-%m")
    try:
        year, mon = int(month[:4]), int(month[5:7])
        if not (1 <= mon <= 12):
            return JSONResponse({"error": "invalid month"}, status_code=400)
        _, last_day = calendar.monthrange(year, mon)
    except Exception:
        return JSONResponse({"error": "invalid month"}, status_code=400)
    stats = _load_scan_stats()
    days = {}
    for d in range(1, last_day + 1):
        key = f"{month}-{d:02d}"
        days[key] = _get_global_scan_count(stats, key)
    return {"month": month, "days": days}

@app.post("/api/scan-stats/reset")
async def reset_scan_stats(request: Request):
    data = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    target_date = data.get("date") or _today_str()
    async with _scan_stats_lock:
        global _scan_stats_cache
        stats = _load_scan_stats()
        _set_global_scan_count(stats, target_date, 0)
        if isinstance(stats.get("devices"), dict):
            for device_id in list(stats["devices"].keys()):
                _set_device_scan_count(stats, device_id, target_date, 0)
        _save_scan_stats(stats)
        _scan_stats_cache = stats
    return {"success": True, "date": target_date, "count": 0, "devices": _device_scan_rows(stats, target_date)}

@app.post("/api/scan-stats/adjust")
async def adjust_scan_stats(request: Request):
    data = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    target_date = data.get("date") or _today_str()
    device_id = _sanitize_device_id(data.get("device_id") or data.get("deviceId"))
    device_label = _sanitize_device_label(data.get("device_label") or data.get("deviceLabel"), device_id)
    delta = data.get("delta")
    try:
        delta = int(delta)
    except (TypeError, ValueError):
        return JSONResponse({"error": "delta must be 1 or -1"}, status_code=400)
    if delta not in (-1, 1):
        return JSONResponse({"error": "delta must be 1 or -1"}, status_code=400)

    async with _scan_stats_lock:
        global _scan_stats_cache
        stats = _load_scan_stats()
        next_count = max(0, _get_global_scan_count(stats, target_date) + delta)
        _set_global_scan_count(stats, target_date, next_count)
        if device_id:
            next_device_count = max(0, _get_device_scan_count(stats, device_id, target_date) + delta)
            _set_device_scan_count(stats, device_id, target_date, next_device_count, device_label)
        _save_scan_stats(stats)
        _scan_stats_cache = stats
    return {"success": True, "date": target_date, "count": next_count, "devices": _device_scan_rows(stats, target_date)}

def _redact_config(cfg: dict) -> dict:
    """返回脱敏副本，密钥字段替换为占位符"""
    import copy
    redacted = copy.deepcopy(cfg)
    for path in _SECRET_FIELDS:
        obj = redacted
        for key in path[:-1]:
            if not isinstance(obj, dict) or key not in obj:
                obj = None
                break
            obj = obj[key]
        if isinstance(obj, dict) and path[-1] in obj and obj[path[-1]]:
            obj[path[-1]] = _REDACTED
    return redacted

def _merge_secrets(existing: dict, incoming: dict) -> dict:
    """合并密钥：incoming 非空且非占位符则更新，否则保留 existing"""
    for path in _SECRET_FIELDS:
        obj_in, obj_ex = incoming, existing
        for key in path[:-1]:
            obj_in = obj_in.get(key, {}) if isinstance(obj_in, dict) else {}
            obj_ex = obj_ex.get(key, {}) if isinstance(obj_ex, dict) else {}
        if not isinstance(obj_in, dict) or not isinstance(obj_ex, dict):
            continue
        field = path[-1]
        new_val = obj_in.get(field, "")
        if not new_val or new_val == _REDACTED:
            # 前端没改密钥 → 保留原值
            obj_in[field] = obj_ex.get(field, "")
    return incoming

@app.get("/api/config")
async def get_config():
    return JSONResponse(_redact_config(get_ocr_config()))

@app.post("/api/config")
async def save_config(request: Request):
    """保存配置（admin面板使用），密钥为空/占位符时保留原值"""
    try:
        incoming = await request.json()
        existing = load_json(CONFIG_FILE, {})
        merged = OcrRuntime._ensure_forward_compat(_merge_secrets(existing, incoming))
        save_json(CONFIG_FILE, merged)
        return {"success": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

@app.get("/api/targets")
async def get_companies():
    return JSONResponse(load_json(TARGETS_FILE, []))

@app.get("/api/template")
async def download_template():
    from openpyxl import Workbook
    from fastapi.responses import FileResponse
    import tempfile
    import os

    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append(['识别对象', '显示信息', '备注'])
    ws.append(['示例科技股份有限公司', '2024-01-15', '示例客户'])
    ws.append(['某某实业集团有限公司', '合同编号: HT-2024-001', ''])

    # 写入临时文件确保浏览器正确识别
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        filename="import_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/api/upload-excel")
async def upload_excel(request: Request):
    data = await request.json()
    file_data = data.get("fileData", "")
    file_name = data.get("fileName", "").lower()
    
    try:
        header, encoded = file_data.split(",", 1)
        file_bytes = base64.b64decode(encoded)
        
        all_rows = []
        if file_name.endswith('.csv'):
            text = file_bytes.decode('utf-8', errors='replace')
            reader = csv.reader(StringIO(text))
            all_rows = list(reader)
        elif file_name.endswith('.xls'):
            return JSONResponse({"error": "不支持的 .xls 格式，请另存为 .xlsx 后再上传"}, status_code=400)
        else:
            wb = openpyxl.load_workbook(filename=BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                # 过滤全空行并转字符串
                str_row = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(str_row):
                    all_rows.append(str_row)
                    
        if not all_rows or len(all_rows) < 2:
            return JSONResponse({"error": "文件为空或缺少数据行"}, status_code=400)
            
        headers = all_rows[0]
        data_rows = all_rows[1:]
        
        name_col = -1
        info_col = -1
        info_col2 = -1
        for i, h in enumerate(headers):
            h_clean = str(h).strip().lower()
            if h_clean in NAME_ALIASES: name_col = i
            elif h_clean in INFO_ALIASES: info_col = i
            elif h_clean in INFO2_ALIASES: info_col2 = i

        preview = []
        for row in data_rows[:5]:
            item = {}
            if name_col >= 0 and name_col < len(row): item['name'] = row[name_col]
            if info_col >= 0 and info_col < len(row): item['displayInfo'] = row[info_col]
            if info_col2 >= 0 and info_col2 < len(row): item['displayInfo2'] = row[info_col2]
            preview.append(item)

        return {
            "headers": headers,
            "preview": preview,
            "totalRows": len(data_rows),
            "nameCol": name_col,
            "infoCol": info_col,
            "infoCol2": info_col2,
            "allRows": data_rows
        }
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

def save_target_headers(headers: List[str], name_col: int, info_col: int, info_col2: int):
    """保存用户上传表格的原始表头到 config.json"""
    cfg = load_json(CONFIG_FILE, {})
    h = {
        "name": headers[name_col] if name_col >= 0 and name_col < len(headers) else "客户名称",
        "displayInfo": headers[info_col] if info_col >= 0 and info_col < len(headers) else "合同总额",
        "displayInfo2": headers[info_col2] if info_col2 >= 0 and info_col2 < len(headers) else "欠款金额"
    }
    if "ui" not in cfg:
        cfg["ui"] = {}
    cfg["ui"]["targetHeaders"] = h
    save_json(CONFIG_FILE, cfg)

@app.post("/api/import-confirm")
async def import_confirm(req: ImportConfirmReq):
    try:
        new_companies = []
        for row in req.allRows:
            name = row[req.nameCol] if req.nameCol < len(row) else ""
            if not name: continue
            info = row[req.infoCol] if (req.infoCol >= 0 and req.infoCol < len(row)) else ""
            info2 = row[req.infoCol2] if (req.infoCol2 >= 0 and req.infoCol2 < len(row)) else ""
            item = {"name": name, "displayInfo": info}
            if info2: item["displayInfo2"] = info2
            new_companies.append(item)

        # 备份原文件
        if os.path.exists(TARGETS_FILE):
            os.rename(TARGETS_FILE, TARGETS_FILE + ".bak")

        save_json(TARGETS_FILE, new_companies)

        # 保存表头映射
        if req.headers:
            save_target_headers(req.headers, req.nameCol, req.infoCol, req.infoCol2)

        return {"success": True, "count": len(new_companies)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


class TargetItem(BaseModel):
    name: str
    displayInfo: str = ""
    displayInfo2: str = ""

class TargetUpdateReq(BaseModel):
    index: int
    name: str
    displayInfo: str = ""
    displayInfo2: str = ""

@app.post("/api/targets/add")
async def add_target(item: TargetItem):
    """添加单条目标"""
    targets = load_json(TARGETS_FILE, [])
    if any(t.get("name") == item.name for t in targets):
        return JSONResponse({"error": "该名称已存在"}, status_code=400)
    targets.append({"name": item.name, "displayInfo": item.displayInfo, "displayInfo2": item.displayInfo2})
    save_json(TARGETS_FILE, targets)
    return {"success": True, "count": len(targets)}

@app.post("/api/targets/update")
async def update_target(req: TargetUpdateReq):
    """编辑指定目标"""
    targets = load_json(TARGETS_FILE, [])
    if req.index < 0 or req.index >= len(targets):
        return JSONResponse({"error": "索引越界"}, status_code=400)
    # 检查重名（排除自身）
    for i, t in enumerate(targets):
        if i != req.index and t.get("name") == req.name:
            return JSONResponse({"error": "该名称已存在"}, status_code=400)
    targets[req.index] = {"name": req.name, "displayInfo": req.displayInfo, "displayInfo2": req.displayInfo2}
    save_json(TARGETS_FILE, targets)
    return {"success": True}

@app.post("/api/targets/delete")
async def delete_target(req: TargetUpdateReq):
    """删除指定目标（只用 index）"""
    targets = load_json(TARGETS_FILE, [])
    if req.index < 0 or req.index >= len(targets):
        return JSONResponse({"error": "索引越界"}, status_code=400)
    targets.pop(req.index)
    save_json(TARGETS_FILE, targets)
    return {"success": True, "count": len(targets)}

@app.post("/api/admin-reset")
async def admin_reset():
    """管理面板强制重置：断开所有 WebSocket 连接，清理状态"""
    disconnected = 0
    for client in list(active_ws_clients):
        try:
            await client.send_json({"status": "reset", "text": "管理员已重置，请重新连接"})
            await client.close()
            disconnected += 1
        except Exception:
            pass
    active_ws_clients.clear()
    return {"success": True, "disconnected": disconnected}

async def _shutdown_process():
    await asyncio.sleep(0.3)
    logger.info("🛑 管理面板请求关闭程序")
    os._exit(0)

@app.post("/api/shutdown")
async def shutdown_app():
    """管理面板关闭程序：响应发送后退出当前 Python 进程"""
    asyncio.create_task(_shutdown_process())
    return {"success": True, "message": "程序正在关闭"}

@app.post("/api/targets/clear")
async def clear_targets():
    """清空所有目标（自动备份）"""
    if os.path.exists(TARGETS_FILE):
        os.rename(TARGETS_FILE, TARGETS_FILE + ".bak")
    save_json(TARGETS_FILE, [])
    return {"success": True, "count": 0}


@app.post("/api/import-merge")
async def import_merge(req: ImportConfirmReq):
    """追加导入，按 name 去重（新数据覆盖旧的 displayInfo）"""
    try:
        # 加载现有数据
        existing = load_json(TARGETS_FILE, [])
        # 用 dict 去重，name 为 key
        merged = {item["name"]: item for item in existing if "name" in item}

        added = 0
        updated = 0
        for row in req.allRows:
            name = row[req.nameCol] if req.nameCol < len(row) else ""
            if not name: continue
            info = row[req.infoCol] if (req.infoCol >= 0 and req.infoCol < len(row)) else ""
            info2 = row[req.infoCol2] if (req.infoCol2 >= 0 and req.infoCol2 < len(row)) else ""
            if name in merged:
                merged[name]["displayInfo"] = info
                if info2: merged[name]["displayInfo2"] = info2
                updated += 1
            else:
                item = {"name": name, "displayInfo": info}
                if info2: item["displayInfo2"] = info2
                merged[name] = item
                added += 1

        # 备份并保存
        if os.path.exists(TARGETS_FILE):
            os.rename(TARGETS_FILE, TARGETS_FILE + ".bak")

        result = list(merged.values())
        save_json(TARGETS_FILE, result)

        # 保存表头映射
        if req.headers:
            save_target_headers(req.headers, req.nameCol, req.infoCol, req.infoCol2)

        return {"success": True, "total": len(result), "added": added, "updated": updated}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/target-headers")
async def get_target_headers():
    """获取当前目标列表的表头标签（从 config.json 读取）"""
    cfg = load_json(CONFIG_FILE, {})
    headers = cfg.get("ui", {}).get("targetHeaders", {
        "name": "客户名称",
        "displayInfo": "合同总额",
        "displayInfo2": "欠款金额"
    })
    return headers

@app.post("/api/open-on-phone")
async def open_on_phone():
    """通过 ADB 在手机上远程打开扫描器页面"""
    try:
        target_serial = _get_adb_target()
        if not target_serial:
            r = subprocess.run(["adb", "devices", "-l"], capture_output=True, text=True, timeout=5)
            return JSONResponse(
                {"status": "error", "message": _adb_error_message_from_devices(_parse_adb_device_lines(r.stdout))},
                status_code=500
            )

        adb_prefix = ["adb", "-s", target_serial]

        # 增强稳定性：先尝试清理旧的映射，再建立新的
        subprocess.run(adb_prefix + ["reverse", "--remove-all"], capture_output=True, timeout=2)
        rev_res = subprocess.run(adb_prefix + ["reverse", f"tcp:{SERVER_PORT}", f"tcp:{SERVER_PORT}"],
                       capture_output=True, text=True, timeout=5)

        if rev_res.returncode != 0:
            logger.error(f"[adb] reverse failed on {target_serial}: {rev_res.stderr}")
            return JSONResponse(
                {"status": "error", "message": rev_res.stderr or rev_res.stdout or "ADB reverse 端口映射失败"},
                status_code=500
            )
        else:
            logger.info(f"[adb] reverse established on {target_serial} for port {SERVER_PORT}")

        # 用 adb shell am start 打开手机浏览器
        target_url = f"http://localhost:{SERVER_PORT}?autostart=1"
        result = subprocess.run(
            adb_prefix + ["shell", "am", "start", "-a", "android.intent.action.VIEW", "-d", target_url],
            capture_output=True, text=True, timeout=5
        )

        if result.returncode == 0:
            return {"status": "success", "message": f"已在手机上打开扫描器 ({target_serial})"}
        else:
            return JSONResponse(
                {"status": "error", "message": result.stderr or "ADB 命令执行失败"},
                status_code=500
            )
    except FileNotFoundError:
        return JSONResponse(
            {"status": "error", "message": "未找到 adb 命令，请确保已安装 Android SDK"},
            status_code=500
        )
    except subprocess.TimeoutExpired:
        return JSONResponse(
            {"status": "error", "message": "ADB 命令超时，请检查手机连接"},
            status_code=500
        )
    except Exception as e:
        return JSONResponse(
            {"status": "error", "message": str(e)},
            status_code=500
        )


# === REST OCR（兜底模式，用于非安全上下文的上传） ===
@app.post("/api/ocr")
async def rest_ocr(request: Request):
    data = await request.json()
    image_data = data.get("image", "")
    device_id = _sanitize_device_id(data.get("device_id") or data.get("deviceId"))
    device_label = _sanitize_device_label(data.get("device_label") or data.get("deviceLabel"), device_id)
    if "," in image_data:
        image_data = image_data.split(",")[1]

    if not image_data:
        return JSONResponse({"error": "No image data"}, status_code=400)

    # REST 端点同样使用信号量保护，防止并发压垮 Ollama
    async with ws_semaphore:
        try:
            text = await asyncio.wait_for(dispatch_ocr(image_data), timeout=60)
        except asyncio.TimeoutError:
            text = "[Timeout: OCR 引擎响应超时，请重试]"
    scan_count = await _maybe_count_scan(text, device_id, device_label)
    return {"text": text, "scan_count": scan_count}

# ================================
# WebSocket 流式引擎（带并发控制）
# ================================
@app.websocket("/ws/ocr")
async def websocket_ocr(websocket: WebSocket):
    device_id = _sanitize_device_id(websocket.query_params.get("device_id") or websocket.query_params.get("deviceId"))
    device_label = _sanitize_device_label(websocket.query_params.get("device_label") or websocket.query_params.get("deviceLabel"), device_id)
    # 并发保护：超过上限时拒绝连接
    if len(active_ws_clients) >= MAX_WS_CONNECTIONS:
        await websocket.accept()
        await websocket.send_json({"status": "error", "text": "服务器繁忙，请稍后重试"})
        await websocket.close()
        logger.info(f"[WS] 拒绝连接：已达上限 {MAX_WS_CONNECTIONS}")
        return

    await websocket.accept()
    active_ws_clients.add(websocket)
    logger.info(f"[WS] Client connected ({len(active_ws_clients)}/{MAX_WS_CONNECTIONS}) device={device_label or device_id or 'unknown'}")

    consecutive_timeouts = 0
    consecutive_empty_frames = 0  # 连续空白/异常小帧计数
    frame_count = 0
    try:
        while True:
            data = await websocket.receive_text()
            frame_count += 1

            # 兼容带有 data:image/jpeg;base64, 前缀的数据
            if "," in data:
                b64_str = data.split(",")[1]
            else:
                b64_str = data

            if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} received, size={len(b64_str)} bytes")

            # 限制图片大小（base64 约 10MB 原始数据）
            if len(b64_str) > 14_000_000:
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} rejected: too large")
                await websocket.send_json({"status": "error", "text": "图片过大"})
                continue

            # 检测明显异常的小帧。窄 ROI 的有效图片可能只有几 KB，阈值不能过高。
            if len(b64_str) < 1200:
                consecutive_empty_frames += 1
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} suspiciously small ({len(b64_str)} bytes), empty streak: {consecutive_empty_frames}")
                if consecutive_empty_frames >= 15:
                    await websocket.send_json({
                        "status": "error",
                        "text": "连续发送空白帧，请检查摄像头是否正常工作或刷新页面重试"
                    })
                    break
                # 发送 processing 让前端保持状态，但继续计数
                await websocket.send_json({"status": "processing", "text": ""})
                continue
            else:
                consecutive_empty_frames = 0

            # 使用信号量限制并发调用，加 60 秒超时防止卡死
            ocr_text = ""
            try:
                async with ws_semaphore:
                    t0 = time.time()
                    ocr_text = await asyncio.wait_for(dispatch_ocr(b64_str), timeout=60)
                    elapsed = int((time.time() - t0) * 1000)
                    logger.info(f"[ocr] frame=#{frame_count} time={elapsed}ms text='{ocr_text[:50]}'")
                consecutive_timeouts = 0
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} OCR result: '{ocr_text[:80]}...' (len={len(ocr_text)})")
            except asyncio.TimeoutError:
                consecutive_timeouts += 1
                ocr_text = "[Timeout: OCR 引擎响应超时，请重试]"
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} OCR timeout")
                if consecutive_timeouts >= 3:
                    await websocket.send_json({
                        "status": "error",
                        "text": "连续多次超时，请刷新页面重试"
                    })
                    break
            except Exception as e:
                ocr_text = f"[OCR Error: {str(e)}]"
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} OCR exception: {e}")

            ocr_text_clean = ocr_text.replace("\n", "").replace(" ", "")

            if len(ocr_text_clean) < 2:
                # 检测结果为空时增加计数，防止无限 processing 循环
                consecutive_empty_frames += 1
                if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} sending 'processing' (text too short: '{ocr_text[:40]}'), empty streak: {consecutive_empty_frames}")
                if consecutive_empty_frames >= 20:
                    await websocket.send_json({
                        "status": "error",
                        "text": "连续识别空白内容，请调整摄像头对准合同文字"
                    })
                    break
                await websocket.send_json({
                    "status": "processing",
                    "text": ocr_text
                })
                continue
            else:
                consecutive_empty_frames = 0

            scan_count = await _maybe_count_scan(ocr_text, device_id, device_label)
            if DEBUG_WS: logger.info(f"[WS] Frame #{frame_count} sending 'success'")
            await websocket.send_json({
                "status": "success",
                "text": ocr_text,
                "scan_count": scan_count
            })

    except WebSocketDisconnect:
        logger.info(f"[WS] Client disconnected ({len(active_ws_clients)}/{MAX_WS_CONNECTIONS})")
    except Exception as e:
        logger.error(f"[WS] Error: {e}")
        try:
            await websocket.close()
        except: pass
    finally:
        active_ws_clients.discard(websocket)

# === 健康检查 ===
@app.get("/health")
async def health_check():
    """检查服务及本地模型状态"""
    cfg = get_ocr_config()
    provider = cfg.get("ocr", {}).get("provider", "mlx")
    mlx_ready = False
    mlx_model = cfg.get("ocr", {}).get("mlx", {}).get("model")
    mlx_message = ""
    if provider == "mlx":
        mlx_ready, mlx_message = _mlx_runtime_status(mlx_model or "mlx-community/GLM-OCR-8bit")

    ollama_ok = False
    ollama_model = None
    ollama_loaded = False
    try:
        ollama_cfg = cfg.get("ocr", {}).get("ollama", {})
        base_url = ollama_cfg.get("baseUrl", "http://localhost:11434").rstrip("/")
        target_model = ollama_cfg.get("model", "glm-ocr")
        ollama_model = target_model
        async with http_session.get(f"{base_url}/api/tags", timeout=aiohttp.ClientTimeout(total=3)) as resp:
            if resp.status == 200:
                data = await resp.json()
                models = [m.get("name", "") for m in data.get("models", [])]
                ollama_ok = len(models) > 0
        async with http_session.get(f"{base_url}/api/ps", timeout=aiohttp.ClientTimeout(total=3)) as resp:
            if resp.status == 200:
                data = await resp.json()
                loaded = [m.get("name", "") for m in data.get("models", [])]
                ollama_loaded = any(target_model in n for n in loaded)
    except Exception:
        pass

    return {
        "status": "ok",
        "provider": provider,
        "mlx": {
            "ready": mlx_ready,
            "model": mlx_model,
            "message": mlx_message
        },
        "ollama": {
            "online": ollama_ok,
            "model": ollama_model,
            "loaded": ollama_loaded
        },
        "active_ws": len(active_ws_clients),
        "max_ws": MAX_WS_CONNECTIONS
    }

def _get_adb_target():
    """获取目标 ADB 设备 serial，优先 USB 设备，解决多设备冲突"""
    r = subprocess.run(["adb", "devices", "-l"], capture_output=True, text=True, timeout=5)
    usb_serial = None
    wifi_serial = None
    for device in _parse_adb_device_lines(r.stdout):
        if device["adb_state"] != "device":
            continue
        if device["mode"] == "usb":
            usb_serial = device["serial"]
        else:
            wifi_serial = device["serial"]
    return usb_serial or wifi_serial

def _get_device_wifi_ip(adb_prefix: list) -> Optional[str]:
    """通过 adb shell ip route 提取手机 WiFi IP"""
    try:
        r = subprocess.run(adb_prefix + ["shell", "ip", "route"],
                           capture_output=True, text=True, timeout=5)
        for line in r.stdout.splitlines():
            if "wlan0" not in line:
                continue
            parts = line.split()
            if "src" in parts:
                idx = parts.index("src")
                if idx + 1 < len(parts):
                    return parts[idx + 1]
            if "via" in parts:
                idx = parts.index("via")
                if idx + 1 < len(parts):
                    return parts[idx + 1]
    except Exception:
        pass
    # 方法2: ip addr（兼容接口名非 wlan0 的设备，如 wlan1 / eth0）
    try:
        r2 = subprocess.run(adb_prefix + ["shell", "ip", "-f", "inet", "addr"],
                            capture_output=True, text=True, timeout=5)
        current_iface = ""
        for line in r2.stdout.splitlines():
            line = line.strip()
            if line and line[0].isdigit() and ":" in line:
                current_iface = line.split(":")[1].strip()
            elif "inet " in line and ("wlan" in current_iface or "eth" in current_iface):
                ip_part = line.split()[1]
                if "/" in ip_part:
                    ip_part = ip_part.split("/")[0]
                if ip_part and not ip_part.startswith("127."):
                    return ip_part
    except Exception:
        pass
    # 方法3: ifconfig fallback（旧 Android / 部分 ROM）
    try:
        r3 = subprocess.run(adb_prefix + ["shell", "ifconfig", "wlan0"],
                            capture_output=True, text=True, timeout=5)
        for line in r3.stdout.splitlines():
            if "inet " in line:
                ip_part = line.split()[1]
                if ":" in ip_part:
                    ip_part = ip_part.split(":")[1]
                return ip_part
    except Exception:
        pass
    return None


def _get_all_adb_devices() -> list:
    """获取所有已连接的 ADB 设备列表，自动去重（当同一手机同时存在 USB 和 WiFi 连接时）"""
    devices_dict = {}
    try:
        r = subprocess.run(["adb", "devices", "-l"], capture_output=True, text=True, timeout=5)
        for device in _parse_adb_device_lines(r.stdout):
            serial = device["serial"]
            mode = device["mode"]
            model = device["model"]
            wifi_ip = None
            adb_prefix = ["adb", "-s", serial]
            if device["adb_state"] == "device":
                try:
                    wifi_ip = _get_device_wifi_ip(adb_prefix)
                except Exception:
                    pass

            device_info = {
                "serial": serial,
                "mode": mode,
                "model": model,
                "wifi_ip": wifi_ip,
                "status": "connected" if device["adb_state"] == "device" else device["adb_state"],
                "adb_state": device["adb_state"]
            }

            # 更加严谨的去重：使用 (model, wifi_ip) 作为联合键
            # 如果同一型号且 IP 相同，判定为同一台手机
            key = (model, wifi_ip) if wifi_ip else (None, serial)

            if key in devices_dict:
                # 关键：优先保留物理 USB 连接，其端口转发最稳定
                if mode == "usb":
                    devices_dict[key] = device_info
            else:
                devices_dict[key] = device_info
                    
    except Exception as e:
        logger.error(f"[adb] enumerate failed: {e}")
    
    return list(devices_dict.values())
def _get_local_ips():
    """获取本机局域网IP地址列表（排除回环地址），带接口名和优先级排序。
    当存在 VPN/隧道/虚拟接口时，优先返回物理网卡（WiFi/以太网）的IP。
    返回: [(ip, iface_name, iface_type, priority), ...]
    """
    PRIVATE_PREFIXES = (
        '192.168.', '10.',
        '172.16.', '172.17.', '172.18.', '172.19.',
        '172.20.', '172.21.', '172.22.', '172.23.',
        '172.24.', '172.25.', '172.26.', '172.27.',
        '172.28.', '172.29.', '172.30.', '172.31.',
    )
    results = []
    seen = set()

    def _iface_priority(name):
        """接口优先级：物理网卡 > 未知 > VPN/虚拟/隧道"""
        low = ('utun', 'tun', 'tap', 'docker', 'veth', 'br-', 'lo', 'gif', 'stf', 'anpi')
        high = ('en0', 'en1', 'en2', 'en3', 'eth0', 'eth1', 'wlan0', 'wlan1', 'wlp')
        if any(name.startswith(p) for p in high):
            return 0
        if any(name.startswith(p) for p in low):
            return 2
        return 1

    def _iface_type(name):
        if name.startswith(('wlan', 'wlp', 'en')):
            return 'wifi'
        if name.startswith(('eth', 'enp')):
            return 'ethernet'
        if name.startswith(('utun', 'tun', 'tap', 'vpn')):
            return 'vpn'
        if any(name.startswith(p) for p in ('docker', 'veth', 'br-', 'lo')):
            return 'virtual'
        return 'unknown'

    def _add(ip, iface):
        if ip in seen or ip == '127.0.0.1':
            return
        if not ip.startswith(PRIVATE_PREFIXES):
            return
        seen.add(ip)
        results.append((ip, iface, _iface_type(iface), _iface_priority(iface)))

    # 方法1: ip addr（现代 Linux）
    try:
        result = subprocess.run(['ip', 'addr'], capture_output=True, text=True)
        current_iface = None
        for line in result.stdout.splitlines():
            line = line.strip()
            if line and not line.startswith('inet'):
                # 例如 "2: en0: <BROADCAST...>"
                m = line.split(':')
                if len(m) >= 2:
                    current_iface = m[1].strip()
            elif 'inet ' in line and current_iface:
                parts = line.split()
                for i, p in enumerate(parts):
                    if p == 'inet' and i + 1 < len(parts):
                        ip = parts[i + 1].split('/')[0]
                        _add(ip, current_iface)
    except Exception:
        pass

    # 方法2: ifconfig（macOS / 旧 Linux）
    if not results:
        try:
            result = subprocess.run(['ifconfig'], capture_output=True, text=True)
            current_iface = None
            for line in result.stdout.splitlines():
                line = line.strip()
                if not line:
                    continue
                iface_part = line.split(':')[0]
                if line[0].isalnum() and ':' in line and ' ' not in iface_part:
                    # 例如 "en0: flags=..."
                    current_iface = iface_part.strip()
                elif 'inet ' in line and current_iface:
                    parts = line.split()
                    for i, p in enumerate(parts):
                        if p == 'inet' and i + 1 < len(parts):
                            ip = parts[i + 1]
                            if ':' in ip:
                                ip = ip.split(':')[1]
                            _add(ip, current_iface)
        except Exception:
            pass

    # 方法3: ipconfig（Windows）
    if not results:
        try:
            result = subprocess.run(['ipconfig'], capture_output=True, text=True)
            current_iface = None
            for line in result.stdout.splitlines():
                line = line.strip()
                if line and line.endswith(':'):
                    current_iface = line.rstrip(':')
                elif 'IPv4 Address' in line and current_iface:
                    ip = line.split(':')[-1].strip()
                    _add(ip, current_iface)
        except Exception:
            pass

    # 方法4: UDP socket 兜底
    if not results:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(0)
            try:
                s.connect(('10.254.254.254', 1))
                ip = s.getsockname()[0]
                _add(ip, 'default')
            except Exception:
                pass
            finally:
                s.close()
        except Exception:
            pass

    # 按优先级排序（物理网卡优先）
    results.sort(key=lambda x: x[3])
    return results


def _tailscale_cmd():
    candidates = [
        "tailscale",
        "/Applications/Tailscale.app/Contents/MacOS/Tailscale",
        "/opt/homebrew/bin/tailscale",
        "/usr/local/bin/tailscale",
    ]
    for cmd in candidates:
        try:
            r = subprocess.run([cmd, "version"], capture_output=True, text=True, timeout=3)
            if r.returncode == 0:
                return cmd
        except Exception:
            continue
    return None


def _run_tailscale(args, timeout=8):
    cmd = _tailscale_cmd()
    if not cmd:
        raise FileNotFoundError("未找到 Tailscale 命令。请在 Tailscale 设置里启用 CLI integration，或确认 Tailscale.app 已安装。")
    return subprocess.run([cmd] + args, capture_output=True, text=True, timeout=timeout)


def _tailscale_state():
    state = {
        "installed": False,
        "connected": False,
        "serve_running": False,
        "dns_name": None,
        "ip": None,
        "url": None,
        "serve_status": "",
        "message": "",
    }
    cmd = _tailscale_cmd()
    if not cmd:
        state["message"] = "未找到 Tailscale 命令"
        return state

    state["installed"] = True
    try:
        r = subprocess.run([cmd, "status", "--json"], capture_output=True, text=True, timeout=5)
        if r.returncode != 0:
            state["message"] = (r.stderr or r.stdout or "Tailscale 未连接").strip()
            return state
        data = json.loads(r.stdout or "{}")
        state["connected"] = data.get("BackendState") == "Running"
        self_info = data.get("Self") or {}
        dns_name = (self_info.get("DNSName") or "").rstrip(".")
        ips = self_info.get("TailscaleIPs") or []
        state["dns_name"] = dns_name or None
        state["ip"] = ips[0] if ips else None
        state["url"] = f"https://{dns_name}/" if dns_name else None
    except Exception as e:
        state["message"] = f"读取 Tailscale 状态失败: {e}"
        return state

    try:
        r = subprocess.run([cmd, "serve", "status"], capture_output=True, text=True, timeout=5)
        state["serve_status"] = (r.stdout or r.stderr or "").strip()
        state["serve_running"] = (
            r.returncode == 0
            and f"http://localhost:{SERVER_PORT}" in state["serve_status"]
            and ("https://" in state["serve_status"] or "proxy" in state["serve_status"])
        )
    except Exception:
        state["serve_running"] = False

    if not state["connected"]:
        state["message"] = "Tailscale 未连接"
    elif not state["url"]:
        state["message"] = "未获取到 MagicDNS 地址，请确认 MagicDNS 已开启"
    elif not state["serve_running"]:
        state["message"] = "Tailscale 已连接，但还未发布扫描服务"
    else:
        state["message"] = "Tailscale 扫描地址已就绪"
    return state


@app.get("/api/network-info")
async def network_info():
    """返回本机局域网IP，供手机WiFi模式连接使用"""
    iface_list = _get_local_ips()  # [(ip, name, type, priority), ...]
    plain_ips = [r[0] for r in iface_list]
    best_ip = plain_ips[0] if plain_ips else None
    return {
        "port": SERVER_PORT,
        "local_ips": plain_ips,
        "interfaces": [
            {"ip": r[0], "name": r[1], "type": r[2]} for r in iface_list
        ],
        "wifi_url": f"http://{best_ip}:{SERVER_PORT}" if best_ip else None,
        "usb_url": f"http://localhost:{SERVER_PORT}"
    }


@app.get("/api/tailscale-status")
async def tailscale_status():
    """返回 Tailscale 连接和 Serve 发布状态。"""
    return _tailscale_state()


@app.post("/api/tailscale-serve-start")
async def tailscale_serve_start():
    """一键启用 Tailscale Serve，把本机扫描服务发布成 HTTPS 地址。"""
    try:
        r = _run_tailscale(["serve", "--bg", f"http://localhost:{SERVER_PORT}"], timeout=12)
        if r.returncode != 0:
            return JSONResponse(
                {"status": "error", "message": (r.stderr or r.stdout or "Tailscale Serve 启动失败").strip()},
                status_code=500,
            )
        state = _tailscale_state()
        state["status"] = "success"
        state["output"] = (r.stdout or r.stderr or "").strip()
        return state
    except FileNotFoundError as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=503)
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)


@app.get("/api/qr-code")
async def qr_code(data: str):
    """服务端生成 QR 码 PNG（避免依赖外部 CDN，支持纯内网环境）"""
    try:
        import qrcode
        from fastapi.responses import StreamingResponse
        img = qrcode.make(data)
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")
    except ImportError:
        return JSONResponse(
            {"error": "qrcode module not installed. Run: pip install qrcode[pil]"},
            status_code=503
        )

@app.get("/api/adb-wifi-status")
async def adb_wifi_status():
    """检查 ADB 连接状态和手机 WiFi IP"""
    result = {"connected": False, "usb_device": None, "wifi_ip": None, "mode": None, "devices": []}
    try:
        r = subprocess.run(["adb", "devices", "-l"], capture_output=True, text=True, timeout=5)
        target_serial = None
        devices = _parse_adb_device_lines(r.stdout)
        result["devices"] = _adb_device_public_fields(devices)
        usable_devices = [d for d in devices if d["adb_state"] == "device"]
        usb_devices = [d for d in usable_devices if d["mode"] == "usb"]
        target_device = usb_devices[0] if usb_devices else (usable_devices[0] if usable_devices else None)
        if target_device:
            target_serial = target_device["serial"]
            if target_device["mode"] == "usb":
                result["usb_device"] = target_serial
                result["mode"] = "usb"
            else:
                result["mode"] = "wifi"
            result["connected"] = True
        else:
            result["message"] = _adb_error_message_from_devices(devices)
        if result["connected"] and target_serial:
            wifi_ip = _get_device_wifi_ip(["adb", "-s", target_serial])
            if wifi_ip:
                result["wifi_ip"] = wifi_ip
    except Exception as e:
        result["error"] = str(e)
    return result


@app.post("/api/adb-wifi-start")
async def adb_wifi_start():
    """开启 ADB 网络模式 (adb tcpip 5555)"""
    try:
        target = _get_adb_target()
        if not target:
            r = subprocess.run(["adb", "devices", "-l"], capture_output=True, text=True, timeout=5)
            return {"status": "error", "message": _adb_error_message_from_devices(_parse_adb_device_lines(r.stdout))}
        adb_prefix = ["adb", "-s", target]
        r = subprocess.run(adb_prefix + ["tcpip", "5555"], capture_output=True, text=True, timeout=10)
        if r.returncode == 0:
            # tcpip 后 ADB daemon 重启，连接短暂断开，等 2 秒再获取 IP
            import time; await asyncio.sleep(2)
            wifi_ip = _get_device_wifi_ip(adb_prefix)
            global adb_wifi_ip_cache
            adb_wifi_ip_cache = wifi_ip
            return {"status": "success", "message": "ADB 网络模式已开启，请拔掉数据线", "wifi_ip": wifi_ip, "port": 5555}
        else:
            return {"status": "error", "message": r.stderr or "adb tcpip 失败"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/adb-wifi-connect")
async def adb_wifi_connect(req: AdbWifiConnectReq):
    """通过 WiFi 连接 ADB 并设置端口映射。
    拔线后 _get_adb_target() 会返回 None，因此不再依赖它检测设备，
    而是直接 trust 前端传来的 wifi_ip，adb connect 后再 reverse。
    """
    wifi_ip = req.wifi_ip
    if not wifi_ip:
        target = _get_adb_target()
        if target:
            wifi_ip = _get_device_wifi_ip(["adb", "-s", target])
    # 兜底：使用 tcpip 时缓存的 IP
    if not wifi_ip and adb_wifi_ip_cache:
        wifi_ip = adb_wifi_ip_cache
    if not wifi_ip:
        return {"status": "error", "message": "缺少手机 WiFi IP，请重新开启无线调试"}

    try:
        # 连接 WiFi ADB（拔线后这是唯一可靠的连接方式）
        wifi_target = f"{wifi_ip}:5555"
        r1 = subprocess.run(["adb", "connect", wifi_target], capture_output=True, text=True, timeout=10)
        out = r1.stdout.lower()
        # 兼容中英文输出: "connected to" / "already connected" / "已连接到"
        if not any(k in out for k in ("connected", "already", "已连接")):
            return {"status": "error", "message": f"连接失败: {r1.stdout or r1.stderr}"}

        # 设置端口映射（使用 WiFi serial，确保 USB 拔掉后仍然有效）
        wifi_prefix = ["adb", "-s", wifi_target]
        r2 = subprocess.run(wifi_prefix + ["reverse", f"tcp:{SERVER_PORT}", f"tcp:{SERVER_PORT}"],
                            capture_output=True, text=True, timeout=5)

        return {
            "status": "success",
            "message": "WiFi ADB 连接成功",
            "wifi_ip": wifi_ip,
            "phone_url": f"http://localhost:{SERVER_PORT}",
            "adb_output": r1.stdout.strip()
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}



# === 多设备 ADB 管理 ===
@app.get("/api/adb-devices")
async def adb_devices_list():
    """列出所有已连接的 ADB 设备"""
    devices = _get_all_adb_devices()
    message = "" if devices else _adb_error_message_from_devices([])
    return {"devices": devices, "count": len(devices), "message": message}


@app.post("/api/adb-repair")
async def adb_repair():
    """重启 ADB server 并重新诊断 USB/授权状态。"""
    steps = []

    def run_adb(cmd: list, timeout: int):
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
            steps.append({
                "cmd": " ".join(cmd),
                "returncode": r.returncode,
                "stdout": (r.stdout or "").strip(),
                "stderr": (r.stderr or "").strip(),
            })
            return r
        except Exception as e:
            steps.append({
                "cmd": " ".join(cmd),
                "returncode": -1,
                "stdout": "",
                "stderr": str(e),
            })
            return None

    run_adb(["adb", "kill-server"], 5)
    await asyncio.sleep(0.5)
    run_adb(["adb", "start-server"], 10)
    await asyncio.sleep(0.5)
    devices_result = run_adb(["adb", "devices", "-l"], 5)
    output = devices_result.stdout if devices_result else ""
    devices = _parse_adb_device_lines(output)
    usable_devices = [d for d in devices if d["adb_state"] == "device"]
    usb_devices = _get_usb_android_devices()
    if usable_devices:
        message = "ADB 已识别 Pixel，可以继续开启无线连接。"
        status = "success"
    else:
        message = _adb_error_message_from_devices(devices)
        status = "needs_phone_action" if usb_devices else "error"
    return {
        "status": status,
        "message": message,
        "devices": _adb_device_public_fields(devices),
        "usb_devices": usb_devices,
        "steps": steps,
    }

@app.post("/api/adb-wifi-start-all")
async def adb_wifi_start_all():
    """批量开启所有 USB 设备的 ADB 网络模式"""
    devices = _get_all_adb_devices()
    usb_devices = [d for d in devices if d["mode"] == "usb" and d.get("adb_state", "device") == "device"]
    if not usb_devices:
        return {"status": "error", "message": "未检测到可用 USB 设备。若 Pixel 显示 unauthorized，请先在手机上允许 USB 调试。"}

    results = []
    for d in usb_devices:
        serial = d["serial"]
        try:
            r = subprocess.run(["adb", "-s", serial, "tcpip", "5555"],
                             capture_output=True, text=True, timeout=10)
            if r.returncode == 0:
                await asyncio.sleep(2)
                wifi_ip = _get_device_wifi_ip(["adb", "-s", serial])
                results.append({"serial": serial, "status": "success", "wifi_ip": wifi_ip})
                logger.info(f"[adb] tcpip 5555 OK for {serial}, wifi_ip={wifi_ip}")
            else:
                results.append({"serial": serial, "status": "error", "message": r.stderr})
                logger.error(f"[adb] tcpip 5555 failed for {serial}: {r.stderr}")
        except Exception as e:
            results.append({"serial": serial, "status": "error", "message": str(e)})
            logger.error(f"[adb] tcpip exception for {serial}: {e}")

    success_count = sum(1 for r in results if r["status"] == "success")
    return {
        "status": "success" if success_count > 0 else "error",
        "message": f"已开启 {success_count}/{len(usb_devices)} 台设备的网络模式",
        "results": results
    }

@app.post("/api/adb-wifi-connect-all")
async def adb_wifi_connect_all():
    """批量连接所有已知 WiFi 设备并设置端口映射"""
    devices = _get_all_adb_devices()
    # 收集所有设备的 WiFi IP（包括已连接的 WiFi 设备和缓存中的 USB 设备 IP）
    targets = []
    for d in devices:
        if d["mode"] == "wifi" and ":" in d["serial"]:
            # 已通过 WiFi 连接的设备，确保 reverse 存在
            targets.append({"serial": d["serial"], "wifi_ip": d["serial"].split(":")[0]})
        elif d["wifi_ip"]:
            # USB 设备但已知 WiFi IP
            targets.append({"serial": d["serial"], "wifi_ip": d["wifi_ip"]})

    if not targets:
        return {"status": "error", "message": "未发现可连接的 WiFi 设备（请先开启无线调试）"}

    results = []
    for t in targets:
        wifi_ip = t["wifi_ip"]
        wifi_target = f"{wifi_ip}:5555"
        try:
            # Connect
            r1 = subprocess.run(["adb", "connect", wifi_target],
                              capture_output=True, text=True, timeout=10)
            out = r1.stdout.lower()
            connected = any(k in out for k in ("connected", "already", "已连接"))
            
            # Reverse
            r2 = subprocess.run(["adb", "-s", wifi_target, "reverse", f"tcp:{SERVER_PORT}", f"tcp:{SERVER_PORT}"],
                              capture_output=True, text=True, timeout=5)
            
            results.append({
                "wifi_ip": wifi_ip,
                "connect": "ok" if connected else r1.stdout.strip(),
                "reverse": "ok" if r2.returncode == 0 else r2.stderr.strip()
            })
            logger.info(f"[adb] connect+reverse {wifi_target}: connect={connected} reverse={r2.returncode==0}")
        except Exception as e:
            results.append({"wifi_ip": wifi_ip, "connect": "error", "message": str(e)})
            logger.error(f"[adb] connect+reverse {wifi_target} failed: {e}")

    ok_count = sum(1 for r in results if r.get("connect") == "ok")
    return {
        "status": "success" if ok_count > 0 else "error",
        "message": f"已连接 {ok_count}/{len(targets)} 台设备",
        "results": results
    }

@app.post("/api/open-on-phone/{serial}")
async def open_on_phone_by_serial(serial: str):
    """在指定设备上打开扫描器页面"""
    try:
        adb_prefix = ["adb", "-s", serial]
        
        # 清理并重建映射
        subprocess.run(adb_prefix + ["reverse", "--remove-all"], capture_output=True, timeout=2)
        rev_res = subprocess.run(adb_prefix + ["reverse", f"tcp:{SERVER_PORT}", f"tcp:{SERVER_PORT}"],
                       capture_output=True, text=True, timeout=5)
        if rev_res.returncode != 0:
            logger.error(f"[adb] reverse failed on {serial}: {rev_res.stderr}")
            return JSONResponse(
                {"status": "error", "message": rev_res.stderr or rev_res.stdout or "ADB reverse 端口映射失败"},
                status_code=500
            )
        
        target_url = f"http://localhost:{SERVER_PORT}?autostart=1"
        result = subprocess.run(
            adb_prefix + ["shell", "am", "start", "-a", "android.intent.action.VIEW", "-d", target_url],
            capture_output=True, text=True, timeout=5
        )
        if result.returncode == 0:
            logger.info(f"[adb] opened browser on {serial}")
            return {"status": "success", "message": f"已在设备 {serial} 上打开扫描器"}
        else:
            return JSONResponse({"status": "error", "message": result.stderr}, status_code=500)
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

@app.get("/api/model-status")
async def model_status():
    """检查当前 OCR 引擎状态"""
    cfg = get_ocr_config()
    provider = cfg["ocr"].get("provider", "mlx")
    result = {"provider": provider, "ready": False, "message": ""}

    try:
        if provider == "mlx":
            mlx_cfg = cfg["ocr"].get("mlx", {})
            model_name = mlx_cfg.get("model", "mlx-community/GLM-OCR-8bit")
            mlx_ready, mlx_message = _mlx_runtime_status(model_name)
            result["ready"] = mlx_ready
            if mlx_ready:
                result["message"] = f"MLX 就绪: {model_name}"
            else:
                paddle_ready, paddle_message = _paddle_runtime_status()
                result["ready"] = paddle_ready
                result["message"] = (
                    f"MLX 未就绪，已降级到 PaddleOCR: {mlx_message}"
                    if paddle_ready
                    else f"MLX 不可用，且 PaddleOCR 不可用: {mlx_message}; {paddle_message}"
                )
        elif provider == "ollama":
            ollama_cfg = cfg["ocr"].get("ollama", {})
            target_model = ollama_cfg.get("model", "glm-ocr")
            base_url = ollama_cfg.get("baseUrl", "http://localhost:11434")
            async with http_session.get(f"{base_url}/api/ps", timeout=aiohttp.ClientTimeout(total=3)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    loaded = [m.get("name", "") for m in data.get("models", [])]
                    if any(target_model in n for n in loaded):
                        result["ready"] = True
                        result["message"] = f"{target_model} loaded"
                    else:
                        result["message"] = "Standby"
        elif provider == "baidu":
            baidu_cfg = cfg["ocr"].get("baidu", {})
            result["ready"] = bool(baidu_cfg.get("apiKey"))
            result["message"] = "已配置" if result["ready"] else "未配置"
        elif provider == "ocrspace":
            ocrspace_cfg = cfg["ocr"].get("ocrspace", {})
            result["ready"] = bool(ocrspace_cfg.get("apiKey"))
            result["message"] = "已配置" if result["ready"] else "未配置"
        elif provider == "openai":
            openai_cfg = cfg["ocr"].get("openai", {})
            result["ready"] = bool(openai_cfg.get("apiKey"))
            result["message"] = "已配置" if result["ready"] else "未配置"
        elif provider == "paddle":
            try:
                import paddleocr
                result["ready"] = True
                result["message"] = f"PaddleOCR {paddleocr.__version__} 就绪"
            except ImportError:
                result["message"] = "未安装 paddleocr"
    except Exception:
        result["message"] = "连接失败"

    return result

@app.post("/api/model-load")
async def model_load():
    """触发当前本地模型加载（异步后台执行）"""
    cfg = get_ocr_config()
    provider = cfg["ocr"].get("provider", "mlx")
    if provider == "mlx":
        asyncio.create_task(_do_mlx_model_load(cfg["ocr"].get("mlx", {})))
        model = cfg["ocr"].get("mlx", {}).get("model", "mlx-community/GLM-OCR-8bit")
        return {"ok": True, "message": f"正在加载 {model}...", "already_loaded": False}
    if provider != "ollama":
        return {"ok": True, "message": "当前引擎无需手动加载"}

    ollama_cfg = cfg["ocr"].get("ollama", {})
    target_model = ollama_cfg.get("model", "glm-ocr")
    base_url = ollama_cfg.get("baseUrl", "http://localhost:11434")
    keep_alive = ollama_cfg.get("keepAlive", "10m")

    # 先检查是否已加载
    try:
        async with http_session.get(f"{base_url}/api/ps", timeout=aiohttp.ClientTimeout(total=3)) as resp:
            if resp.status == 200:
                data = await resp.json()
                loaded = [m.get("name", "") for m in data.get("models", [])]
                if any(target_model in n for n in loaded):
                    return {"ok": True, "message": f"{target_model} 已在内存中", "already_loaded": True}
    except Exception:
        pass

    # 后台启动加载
    asyncio.create_task(_do_model_load(base_url, target_model, keep_alive))
    return {"ok": True, "message": f"正在加载 {target_model}...", "already_loaded": False}

async def _do_model_load(base_url: str, model: str, keep_alive: str):
    """后台加载模型到 GPU/内存"""
    timeout = aiohttp.ClientTimeout(total=120, sock_read=30, connect=10)
    try:
        url = f"{base_url.rstrip('/')}/api/generate"
        payload = {
            "model": model,
            "prompt": "hi",
            "stream": False,
            "keep_alive": keep_alive,
            "options": {"num_predict": 1, "temperature": 0},
        }
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.post(url, json=payload, timeout=timeout) as resp:
                if resp.status == 200:
                    logger.info(f"✅ 模型 {model} 加载完成")
                else:
                    logger.warning(f"⚠️ 模型加载失败: HTTP {resp.status}")
    except asyncio.TimeoutError:
        logger.warning(f"⚠️ 模型加载超时，模型可能已卡死")
    except Exception as e:
        logger.warning(f"⚠️ 模型加载异常: {e}")

async def _do_mlx_model_load(mlx_cfg: dict):
    """后台加载 MLX 模型到内存。"""
    try:
        text = await _runtime_for_ocr().recognize_mlx(_warmup_image_b64(), mlx_cfg)
        if text.startswith("[MLX Error:"):
            logger.warning(f"⚠️ MLX 模型加载失败: {text}")
        else:
            logger.info("✅ MLX 模型加载完成")
    except Exception as e:
        logger.warning(f"⚠️ MLX 模型加载异常: {e}")

# === 日志 API ===
@app.get("/api/logs")
async def get_logs(lines: int = 100):
    """读取最近 N 行日志"""
    try:
        if not os.path.exists(LOG_FILE):
            return {"logs": "", "lines": 0}
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            all_lines = f.readlines()
        recent = all_lines[-lines:]
        return {"logs": "".join(recent), "lines": len(recent)}
    except Exception as e:
        return {"logs": f"Error reading logs: {e}", "lines": 0}

@app.get("/api/logs/download")
async def download_logs():
    """下载完整日志文件"""
    if os.path.exists(LOG_FILE):
        return FileResponse(LOG_FILE, filename="scanner.log", media_type="text/plain")
    return JSONResponse({"error": "Log file not found"}, status_code=404)

# === 根路由（显式声明，优先于 StaticFiles mount） ===
@app.get("/")
async def root():
    return FileResponse(
        os.path.join(BASE_DIR, "index.html"),
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        }
    )

# === 静态文件（禁用 HTML/JS/CSS 缓存，防止手机浏览器缓存旧版本） ===
class NoCacheStaticFiles(StaticFiles):
    async def get_response(self, path: str, scope):
        response = await super().get_response(path, scope)
        if path.endswith((".html", ".js", ".css")):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response

app.mount("/", NoCacheStaticFiles(directory=BASE_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    logger.info(f"🚀 FastAPI Server starting on http://localhost:{SERVER_PORT}")
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=SERVER_PORT,
        log_level="info",
        ws="websockets-sansio",
        ws_ping_interval=None,
        ws_per_message_deflate=False,
    )
